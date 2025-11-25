import datetime
import io
import uuid

import openpyxl
from openpyxl.styles import numbers

from core.tests_base.test_admin import TestAdminBase
from travels import models


class ClientAdminTestCase(TestAdminBase):
    """Testing client admin"""

    def setUp(self):
        super().setUp()
        self.endpoint = "/admin/travels/client/"

    def test_search_bar(self):
        """Validate search bar working"""

        self.submit_search_bar(self.endpoint)


class VipCodeAdminTestCase(TestAdminBase):
    """Testing code admin"""

    def setUp(self):
        super().setUp()
        self.endpoint = "/admin/travels/vipcode/"

    def test_search_bar(self):
        """Validate search bar working"""

        self.submit_search_bar(self.endpoint)


class VehicleAdminTestCase(TestAdminBase):
    """Testing vehicle admin"""

    def setUp(self):
        super().setUp()
        self.endpoint = "/admin/travels/vehicle/"


class SaleAdminTestCase(TestAdminBase):
    """Testing sale admin"""

    def setUp(self):
        super().setUp()
        self.endpoint = "/admin/travels/sale/"
        random_suffix = uuid.uuid4().hex
        self.export_zone = models.Zone.objects.create(
            name=f"Export zone {random_suffix}"
        )
        self.export_location = models.Location.objects.create(
            name=f"Export location {random_suffix}", zone=self.export_zone
        )
        self.export_service_type = models.ServiceType.objects.create(
            name=f"Export service {random_suffix}"
        )
        self.export_vehicle = models.Vehicle.objects.create(
            name=f"Export vehicle {random_suffix}"
        )

    def _create_export_sale(
        self,
        *,
        client_name: str,
        client_last_name: str,
        email: str,
        phone: str,
        passengers: int,
        arrival_date: datetime.date,
        arrival_time: datetime.time,
        arrival_airline: str,
        arrival_flight: str,
        departure_date: datetime.date,
        departure_time: datetime.time,
        departure_airline: str,
        departure_flight: str,
    ):
        client = models.Client.objects.create(
            name=client_name,
            last_name=client_last_name,
            email=email,
            phone=phone,
        )

        sale = models.Sale.objects.create(
            client=client,
            vip_code=None,
            vehicle=self.export_vehicle,
            passengers=passengers,
            service_type=self.export_service_type,
            location=self.export_location,
            total=250.0,
            paid=True,
        )

        arrival = models.Transfer.objects.create(
            date=arrival_date,
            hour=arrival_time,
            type="arrival",
            sale=sale,
            airline=arrival_airline,
            flight_number=arrival_flight,
        )

        departure = models.Transfer.objects.create(
            date=departure_date,
            hour=departure_time,
            type="departure",
            sale=sale,
            airline=departure_airline,
            flight_number=departure_flight,
        )

        return sale, client, arrival, departure

    def _run_export_action(self, sale_ids: list[int]):
        response = self.client.post(
            self.endpoint,
            {
                "action": "export_to_excel",
                "_selected_action": [str(sale_id) for sale_id in sale_ids],
            },
        )

        self.assertEqual(
            response.status_code,
            200,
            msg="Export action should return the generated workbook",
        )
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        content = b"".join(response.streaming_content)
        return openpyxl.load_workbook(io.BytesIO(content))

    def _normalize_export_row(self, row: list):
        date_columns = {8, 12}
        normalized_row = []
        for index, value in enumerate(row, start=1):
            if index in date_columns and isinstance(value, datetime.datetime):
                normalized_row.append(value.date())
            elif value is None:
                # Convert None to empty string for consistency
                normalized_row.append("")
            else:
                normalized_row.append(value)
        return normalized_row

    def test_search_bar(self):
        """Validate search bar working"""

        self.submit_search_bar(self.endpoint)

    def test_custom_links(self):
        """Validate custom custom links"""

        links = {
            "Transportaciones": "/admin/travels/transfer/",
        }

        # Open employee list page
        response = self.client.get("/admin/travels/sale/")

        # Validate links
        for link_text, link in links.items():
            self.assertContains(response, link_text)
            self.assertContains(response, link)

    def test_export_to_excel_single_row(self):
        """Ensure export action writes a single sale starting at row 6"""

        sale, client, arrival, departure = self._create_export_sale(
            client_name="Exportable",
            client_last_name="Client",
            email="export@example.com",
            phone="5550001234",
            passengers=3,
            arrival_date=datetime.date(2025, 12, 1),
            arrival_time=datetime.time(9, 45),
            arrival_airline="ArrivalAir",
            arrival_flight="ARR123",
            departure_date=datetime.date(2025, 12, 2),
            departure_time=datetime.time(18, 30),
            departure_airline="DepartAir",
            departure_flight="DEP456",
        )

        workbook = self._run_export_action([sale.id])
        sheet = workbook.active

        expected_headers = [
            "Last Name",
            "First Name",
            "email",
            "Phone",
            "Hotel",
            "Vehicle ",
            "Number of Guests",
            "Arrival Date",
            "Airline",
            "Flight",
            "Arrival Time",
            "Departure Date",
            "Airline",
            "Flight",
            "Departure Time",
        ]
        actual_headers = [
            sheet.cell(row=5, column=index).value for index in range(1, 16)
        ]
        self.assertEqual(expected_headers, actual_headers)

        row_values = [sheet.cell(row=6, column=index).value for index in range(1, 16)]
        row_values = self._normalize_export_row(row_values)
        self.assertEqual(
            row_values,
            [
                client.last_name,
                client.name,
                client.email,
                client.phone,
                self.export_location.name,  # Hotel
                self.export_vehicle.name,
                sale.passengers,
                arrival.date,
                arrival.airline,
                arrival.flight_number,
                arrival.hour,
                departure.date,
                departure.airline,
                departure.flight_number,
                departure.hour,
            ],
        )

    def test_export_to_excel_multiple_rows(self):
        """Ensure selected sales produce consecutive rows in the export"""

        sale1, client1, arrival1, departure1 = self._create_export_sale(
            client_name="First",
            client_last_name="Row",
            email="first@example.com",
            phone="1110001234",
            passengers=2,
            arrival_date=datetime.date(2025, 12, 3),
            arrival_time=datetime.time(8, 15),
            arrival_airline="ArrivalOne",
            arrival_flight="ONE123",
            departure_date=datetime.date(2025, 12, 4),
            departure_time=datetime.time(19, 45),
            departure_airline="DepartOne",
            departure_flight="ONE456",
        )

        sale2, client2, arrival2, departure2 = self._create_export_sale(
            client_name="Second",
            client_last_name="Row",
            email="second@example.com",
            phone="2220001234",
            passengers=4,
            arrival_date=datetime.date(2025, 12, 5),
            arrival_time=datetime.time(10, 0),
            arrival_airline="ArrivalTwo",
            arrival_flight="TWO123",
            departure_date=datetime.date(2025, 12, 6),
            departure_time=datetime.time(20, 15),
            departure_airline="DepartTwo",
            departure_flight="TWO456",
        )

        workbook = self._run_export_action([sale1.id, sale2.id])
        sheet = workbook.active

        rows = []
        for row_index in (6, 7):
            row_values = [sheet.cell(row=row_index, column=col).value for col in range(1, 16)]
            rows.append(self._normalize_export_row(row_values))

        expected_rows = [
            [
                client1.last_name,
                client1.name,
                client1.email,
                client1.phone,
                self.export_location.name,  # Hotel
                self.export_vehicle.name,
                sale1.passengers,
                arrival1.date,
                arrival1.airline,
                arrival1.flight_number,
                arrival1.hour,
                departure1.date,
                departure1.airline,
                departure1.flight_number,
                departure1.hour,
            ],
            [
                client2.last_name,
                client2.name,
                client2.email,
                client2.phone,
                self.export_location.name,  # Hotel
                self.export_vehicle.name,
                sale2.passengers,
                arrival2.date,
                arrival2.airline,
                arrival2.flight_number,
                arrival2.hour,
                departure2.date,
                departure2.airline,
                departure2.flight_number,
                departure2.hour,
            ],
        ]

        self.assertCountEqual(expected_rows, rows)

    def test_export_date_time_number_formats_applied(self):
        """Ensure date/time cells use the configured number formats"""

        sale, _, arrival, departure = self._create_export_sale(
            client_name="Formatting",
            client_last_name="Check",
            email="format@example.com",
            phone="5559871234",
            passengers=1,
            arrival_date=datetime.date(2025, 12, 7),
            arrival_time=datetime.time(7, 5),
            arrival_airline="FormatAir",
            arrival_flight="FMT007",
            departure_date=datetime.date(2025, 12, 8),
            departure_time=datetime.time(14, 20),
            departure_airline="FormatDepart",
            departure_flight="FMT008",
        )

        workbook = self._run_export_action([sale.id])
        sheet = workbook.active

        format_expectations = {
            8: numbers.FORMAT_DATE_YYYYMMDD2,
            11: numbers.FORMAT_DATE_TIME4,
            12: numbers.FORMAT_DATE_YYYYMMDD2,
            15: numbers.FORMAT_DATE_TIME4,
        }

        for column, expected_format in format_expectations.items():
            cell = sheet.cell(row=6, column=column)
            self.assertEqual(
                cell.number_format,
                expected_format,
                msg=f"Column {column} should use {expected_format}",
            )
            if column in {8, 12}:
                cell_value = cell.value
                if isinstance(cell_value, datetime.datetime):
                    cell_value = cell_value.date()
                self.assertEqual(
                    cell_value,
                    arrival.date if column == 8 else departure.date,
                )
            else:
                self.assertEqual(
                    cell.value,
                    arrival.hour if column == 11 else departure.hour,
                )


class ServiceTypeAdminTestCase(TestAdminBase):
    """Testing transfer type admin"""

    def setUp(self):
        super().setUp()
        self.endpoint = "/admin/travels/servicetype/"


class TransferAdminTestCase(TestAdminBase):
    """Testing transfer admin"""

    def setUp(self):
        super().setUp()
        self.endpoint = "/admin/travels/transfer/"

    def test_search_bar(self):
        """Validate search bar working"""

        self.submit_search_bar(self.endpoint)


class ZoneAdminTestCase(TestAdminBase):
    """Testing zone admin"""

    def setUp(self):
        super().setUp()
        self.endpoint = "/admin/travels/zone/"

    def test_search_bar(self):
        """Validate search bar working"""

        self.submit_search_bar(self.endpoint)


class LocationAdminTestCase(TestAdminBase):
    """Testing location admin"""

    def setUp(self):
        super().setUp()
        self.endpoint = "/admin/travels/location/"

    def test_search_bar(self):
        """Validate search bar working"""

        self.submit_search_bar(self.endpoint)


class PricingAdminTestCase(TestAdminBase):
    """Testing pricing admin"""

    def setUp(self):
        super().setUp()
        self.endpoint = "/admin/travels/pricing/"

    def test_search_bar(self):
        """Validate search bar working"""

        self.submit_search_bar(self.endpoint)
