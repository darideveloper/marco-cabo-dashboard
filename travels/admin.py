import io
import os

import openpyxl
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter

from django.conf import settings
from django.contrib import admin, messages
from django.http import FileResponse
from django.utils import timezone
from django.utils.html import format_html

from travels import models


@admin.register(models.Zone)
class ZoneAdmin(admin.ModelAdmin):
    list_display = ("name", "created_at")
    list_filter = ("created_at", "updated_at")
    search_fields = ("name",)
    readonly_fields = ("created_at", "updated_at")
    ordering = ("name",)


@admin.register(models.Location)
class LocationAdmin(admin.ModelAdmin):
    list_display = ("name", "zone", "created_at")
    list_filter = ("zone", "created_at", "updated_at")
    search_fields = ("name",)
    readonly_fields = ("created_at", "updated_at")
    ordering = ("name",)


@admin.register(models.Client)
class ClientAdmin(admin.ModelAdmin):
    list_display = ("name", "last_name", "email", "phone", "created_at")
    list_filter = ("created_at", "updated_at")
    search_fields = ("name", "last_name", "email", "phone")
    readonly_fields = ("created_at", "updated_at")
    ordering = ("name",)


@admin.register(models.VipCode)
class VipCodeAdmin(admin.ModelAdmin):
    list_display = ("value", "active", "created_at")
    list_filter = ("active", "created_at", "updated_at")
    search_fields = ("value",)
    readonly_fields = ("created_at", "updated_at")
    ordering = ("value",)


@admin.register(models.Vehicle)
class VehicleAdmin(admin.ModelAdmin):
    list_display = ("name", "created_at")
    list_filter = ("created_at", "updated_at")
    readonly_fields = ("created_at", "updated_at")
    ordering = ("name",)


@admin.register(models.Sale)
class SaleAdmin(admin.ModelAdmin):
    actions = ("export_to_excel",)
    list_display = (
        "stripe_code",
        "client",
        "vip_code",
        "vehicle",
        "service_type",
        "location",
        "passengers",
        "total",
        "paid",
        "custom_links",
        "created_at",
        "updated_at",
    )
    list_filter = (
        "client",
        "vip_code",
        "vehicle",
        "service_type",
        "location",
        "passengers",
        "paid",
        "created_at",
    )
    search_fields = (
        "client__name",
        "client__email",
        "vehicle__name",
        "vip_code__value",
        "total",
    )
    readonly_fields = ("stripe_code", "created_at", "updated_at")
    ordering = ("-created_at",)
    
    # CUSTOM FIELDS
    def custom_links(self, obj):
        """Create custom Imprimir and Ver buttons"""
        return format_html(
            '<a class="btn btn-secondary my-1" href="{}">Transportaciones</a>',
            f"/admin/travels/transfer/?sale__id__exact={obj.id}&q=",
        )
    
    def export_to_excel(self, request, queryset):
        """Export the selected sales to Excel"""

        date_columns = {8, 12}
        time_columns = {11, 15}

        if not queryset:
            self.message_user(
                request,
                "Seleccione al menos una venta para exportar.",
                level=messages.INFO,
            )
            return

        template_path = os.path.join(settings.BASE_DIR, "utils", "export-template.xlsx")
        if not os.path.exists(template_path):
            self.message_user(
                request,
                "No se encontró la plantilla de exportación.",
                level=messages.ERROR,
            )
            return

        workbook = openpyxl.load_workbook(template_path)
        sheet = workbook.active
        queryset = queryset.select_related("client", "vehicle").prefetch_related(
            "transfer_set"
        )

        current_row = 6
        for sale in queryset:
            arrival = sale.transfer_set.filter(type="arrival").first()
            departure = sale.transfer_set.filter(type="departure").first()

            row = [
                sale.client.last_name,
                sale.client.name,
                sale.client.email,
                sale.client.phone,
                sale.location.name,  # Hotel
                sale.vehicle.name,
                sale.passengers,
                arrival.date if arrival else None,
                arrival.airline if arrival else "",
                arrival.flight_number if arrival else "",
                arrival.hour if arrival else None,
                departure.date if departure else None,
                departure.airline if departure else "",
                departure.flight_number if departure else "",
                departure.hour if departure else None,
            ]

            for col_index, value in enumerate(row, start=1):
                cell = sheet.cell(row=current_row, column=col_index, value=value)
                if value is not None:
                    if col_index in date_columns:
                        cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2
                    elif col_index in time_columns:
                        cell.number_format = numbers.FORMAT_DATE_TIME4
            current_row += 1

        for col in range(1, sheet.max_column + 1):
            max_length = 0
            for row in range(1, sheet.max_row + 1):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value is None:
                    continue
                max_length = max(max_length, len(str(cell_value)))
            if max_length:
                sheet.column_dimensions[get_column_letter(col)].width = max_length * 2

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = (
            "marco-cabo-transportaciones-"
            f"{timezone.localtime().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        response = FileResponse(
            output,
            as_attachment=True,
            filename=filename,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        return response
    
    # Labels for custom fields
    custom_links.short_description = "Acciones"
    export_to_excel.short_description = "Exportar a Excel"


@admin.register(models.ServiceType)
class ServiceTypeAdmin(admin.ModelAdmin):
    list_display = ("name", "created_at")
    list_filter = ("created_at", "updated_at")
    readonly_fields = ("created_at", "updated_at")
    ordering = ("name",)


@admin.register(models.Transfer)
class TransferAdmin(admin.ModelAdmin):
    list_display = ("date", "hour", "type", "sale", "created_at")
    list_filter = (
        "type",
        "sale",
        "sale__client",
        "sale__vehicle",
        "sale__location",
        "created_at",
        "updated_at",
    )
    search_fields = (
        "sale__location__name",
        "sale__client__name",
        "sale__client__email",
        "sale__vehicle__name",
        "sale__vip_code__value",
    )
    readonly_fields = ("created_at", "updated_at")
    ordering = ("-created_at",)


@admin.register(models.Pricing)
class PricingAdmin(admin.ModelAdmin):
    list_display = (
        "location",
        "vehicle",
        "service_type",
        "price",
        "created_at",
    )
    list_filter = ("location", "vehicle", "service_type", "created_at", "updated_at")
    search_fields = (
        "location__name",
        "vehicle__name",
        "service_type__name",
        "price",
    )
    readonly_fields = ("created_at", "updated_at")
    ordering = ("location__name", "vehicle__name", "service_type__name")